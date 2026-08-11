import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

folder = r'd:\Dokumen\02_Kerja_Profesional\Juragan by Anak Bawang\Marketplace_Templates'

def audit_file(file_path, expected_skus, expected_prices):
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Template']
    passed = True
    print(f'=== AUDIT {os.path.basename(file_path)} ===')
    for idx, (exp_sku, exp_price) in enumerate(zip(expected_skus, expected_prices)):
        r = 6 + idx
        sku = ws.cell(r, 27).value
        name = ws.cell(r, 3).value
        price = ws.cell(r, 24).value
        stock = ws.cell(r, 26).value
        desc = ws.cell(r, 4).value
        
        errs = []
        if sku != exp_sku: errs.append(f'SKU: {sku} != {exp_sku}')
        if price != exp_price: errs.append(f'Price: {price} != {exp_price}')
        if stock != 50: errs.append(f'Stock: {stock}')
        if not desc or 'HALAL CERTIFIED' not in desc: errs.append('Desc Halal Missing')
        
        if errs:
            passed = False
            print(f'  ❌ Baris {r} ({exp_sku}): {errs}')
        else:
            print(f'  ✅ Baris {r} ({exp_sku}): OK | Harga System = Rp {price:,} | Stock = {stock}')
    print(f'HASIL: {"PASSED ALL CHECKS" if passed else "FAILED"}\n')
    return passed

murni_skus = ['JBM-150', 'JBM-250', 'JBM-500', 'JBM-1K', 'JBM-PAKET2X250', 'JBM-COMBO150-250', 'JBM-PAKETGROSIR1KG', 'JBM-100-TRIAL', 'JBM-HORECA-2KG']
murni_prices = [59000, 89500, 145000, 239000, 179000, 148500, 289000, 39900, 478000]

a_skus = ['JBA-150', 'JBA-250', 'JBA-500', 'JBA-1K', 'JBA-PAKET2X250', 'JBA-COMBO150-250', 'JBA-PAKETGROSIR1KG', 'JBA-100-TRIAL', 'JBA-HORECA-2KG']
a_prices = [49900, 69900, 119000, 199000, 139900, 119800, 238000, 34900, 399000]

audit_file(os.path.join(folder, 'Template_Mass_Upload_Grade_Murni.xlsx'), murni_skus, murni_prices)
audit_file(os.path.join(folder, 'Template_Mass_Upload_Grade_A.xlsx'), a_skus, a_prices)

wb_disc = openpyxl.load_workbook(os.path.join(folder, 'Product_Discount_Shopee_TikTok.xlsx'))
ws_disc = wb_disc.active
print('=== AUDIT Product_Discount_Shopee_TikTok.xlsx ===')
for r in range(2, 20):
    sku = ws_disc.cell(r, 1).value
    p_norm = ws_disc.cell(r, 3).value
    p_promo = ws_disc.cell(r, 4).value
    disc = ws_disc.cell(r, 5).value
    print(f'  Row {r:2}: SKU {sku:<18} | Coret: Rp {p_norm:>8,} | Promo: Rp {p_promo:>8,} | Disc: {disc:>5}%')
